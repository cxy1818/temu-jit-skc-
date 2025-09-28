import sys, os, json, glob, shutil, time, subprocess
from PySide6.QtWidgets import (
    QApplication, QMessageBox, QInputDialog, QFileDialog, QProgressDialog, QTableWidgetItem
)
from PySide6.QtCore import Qt, QThread, Signal
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from ui_manager import SKCUI, status_options

MAX_FILES = 100
projects = {}            
current_project = None

class ExcelSaveThread(QThread):
    progress = Signal(int)
    finished = Signal(str)
    error = Signal(str)

    def __init__(self, project_name):
        super().__init__()
        self.project_name = project_name

    def run(self):
        try:
            project = projects[self.project_name]
            folder = project["folder"]
            db = project["database"]
            imgs = project["images"]

            if not os.path.exists(folder):
                os.makedirs(folder, exist_ok=True)

            files = sorted(glob.glob(os.path.join(folder, "skc_*.xlsx")), key=os.path.getmtime)
            while len(files) >= MAX_FILES:
                try:
                    os.remove(files[0])
                except:
                    pass
                files.pop(0)

            timestamp = time.strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(folder, f"skc_{timestamp}.xlsx")

            wb = Workbook()
            ws = wb.active
            col = 1
            total = max(len(db), 1)
            for i, (product, skcs) in enumerate(db.items(), 1):
                
                try:
                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                    ws.cell(row=1, column=col, value=product)
                except Exception:
                    ws.cell(row=1, column=col, value=product)

               
                if product in imgs and imgs[product]:
                    try:
                        img = XLImage(imgs[product])
                        img.width = 100
                        img.height = 100
                        ws.add_image(img, f"{get_column_letter(col)}2")
                        ws.row_dimensions[2].height = 80
                        ws.column_dimensions[get_column_letter(col)].width = 15
                    except Exception:
                        pass

                ws.cell(row=3, column=col, value="SKC")
                ws.cell(row=3, column=col+1, value="状态")
                r_index = 4
                for skc, status in skcs.items():
                    ws.cell(row=r_index, column=col, value=skc)
                    ws.cell(row=r_index, column=col+1, value=status)
                    r_index += 1
                col += 2

                
                percent = int(i / total * 100)
                self.progress.emit(percent)

            wb.save(filename)
            self.finished.emit(filename)
        except Exception as e:
            self.error.emit(str(e))


def save_project_data(project_name):
    project = projects[project_name]
    folder = project["folder"]
    if not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    data_file = os.path.join(folder, "data.json")
    try:
        with open(data_file, "w", encoding="utf-8") as f:
            json.dump({"database": project["database"], "images": project["images"]}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        QMessageBox.warning(None, "保存失败", f"保存 data.json 失败: {e}")

def load_project_data(project_name):
    project = projects[project_name]
    data_file = os.path.join(project["folder"], "data.json")
    if os.path.exists(data_file):
        try:
            with open(data_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            project["database"] = data.get("database", {})
            project["images"] = data.get("images", {})
        except Exception:
            pass

def load_all_projects():
    global projects, current_project
    folders = sorted([d for d in os.listdir() if os.path.isdir(d) and d.startswith("skc-data-")])
    projects.clear()
    for i, folder in enumerate(folders, 1):
        name = f"项目{i}"
        projects[name] = {"database": {}, "images": {}, "folder": folder}
        load_project_data(name)
    if not projects:
        name = "项目1"
        folder = "skc-data-1"
        os.makedirs(folder, exist_ok=True)
        projects[name] = {"database": {}, "images": {}, "folder": folder}
        save_project_data(name)
    current_project = list(projects.keys())[0]

def get_current_database():
    if current_project and current_project in projects:
        return projects[current_project]["database"], projects[current_project]["images"]
    return None, None

def open_latest_file(project_name):
    folder = projects[project_name]["folder"]
    files = sorted(glob.glob(os.path.join(folder, "skc_*.xlsx")), key=os.path.getmtime)
    if not files:
        QMessageBox.information(None, "提示", "当前没有生成的 Excel 文件")
        return
    latest_file = files[-1]
    try:
        if sys.platform.startswith("win"):
            os.startfile(latest_file)
        elif sys.platform == "darwin":
            subprocess.call(["open", latest_file])
        else:
            subprocess.call(["xdg-open", latest_file])
    except Exception as e:
        QMessageBox.warning(None, "打开失败", f"无法打开文件: {e}")


class SKCManagerLogic:
    def __init__(self, ui: SKCUI):
        self.ui = ui
        self.thread = None
        self.progress_dialog = None
        self.setup()

    def setup(self):
    
        self.ui.btn_add.clicked.connect(self.add_product)
        self.ui.btn_batch_modify.clicked.connect(self.batch_modify_skc)
        self.ui.btn_batch_delete.clicked.connect(self.batch_delete_skc)
        self.ui.btn_auto_sort.clicked.connect(self.auto_sort_by_status)
        self.ui.btn_add_image.clicked.connect(self.confirm_add_image)
        self.ui.btn_import_excel.clicked.connect(self.import_excel_data)
        self.ui.btn_open_latest.clicked.connect(self.open_latest_excel)
        self.ui.btn_clear_skc.clicked.connect(lambda: self.ui.entry_skc.clear())
        self.ui.project_combo.currentTextChanged.connect(self.on_project_changed)
        self.ui.btn_new_project.clicked.connect(self.create_project_ui)
        self.ui.btn_switch_project.clicked.connect(self.switch_project_ui)
        self.ui.btn_import_project.clicked.connect(self.import_project_ui)
        self.ui.btn_export_project.clicked.connect(self.export_project_ui)

        load_all_projects()
        self.ui.project_combo.clear()
        for p in projects.keys():
            self.ui.project_combo.addItem(p)
        if current_project:
            self.ui.project_combo.setCurrentText(current_project)
        self.refresh_table()

    
    def save_database_async(self):
        if not current_project:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        self.progress_dialog = QProgressDialog("正在保存 Excel...", "取消", 0, 100, self.ui)
        self.progress_dialog.setWindowTitle("保存中")
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.show()

        self.thread = ExcelSaveThread(current_project)
        self.thread.progress.connect(self.progress_dialog.setValue)
        self.thread.finished.connect(self.on_save_finished)
        self.thread.error.connect(self.on_save_error)
        self.progress_dialog.canceled.connect(self.thread.terminate)
        self.thread.start()

    def on_save_finished(self, filename):
        if self.progress_dialog:
            self.progress_dialog.close()
        QMessageBox.information(self.ui, "完成", f"已保存: {filename}")

    def on_save_error(self, msg):
        if self.progress_dialog:
            self.progress_dialog.close()
        QMessageBox.warning(self.ui, "保存失败", f"保存失败: {msg}")

    
    def refresh_table(self):
        db, _ = get_current_database()
        if not db:
            self.ui.table.setRowCount(0)
            self.ui.image_product_combo.clear()
            return
        self.ui.image_product_combo.clear()
        self.ui.image_product_combo.addItems(list(db.keys()))
        items = [(p, skc, s) for p, skcs in db.items() for skc, s in skcs.items()]
        self.ui.table.setRowCount(len(items))
        for r, (product, skc, status) in enumerate(items):
            self.ui.table.setItem(r, 0, QTableWidgetItem(str(product)))
            self.ui.table.setItem(r, 1, QTableWidgetItem(str(skc)))
            self.ui.table.setItem(r, 2, QTableWidgetItem(str(status)))

    
    def add_product(self):
        db, _ = get_current_database()
        if db is None:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        product = self.ui.entry_product.text().strip()
        if not product:
            QMessageBox.warning(self.ui, "提示", "请输入货号")
            return
        skc_text = self.ui.entry_skc.text().strip()
        if not skc_text:
            QMessageBox.warning(self.ui, "提示", "请输入 SKC")
            return
        status = self.ui.status_combo.currentText()
        if product not in db:
            db[product] = {}
        added = 0
        for skc in skc_text.split():
            s = str(skc).strip()
            if any(s in skcs for p, skcs in db.items()):
                continue
            db[product][s] = status
            added += 1
        save_project_data(current_project)
        self.save_database_async()
        QMessageBox.information(self.ui, "完成", f"新增 {added} 个 SKC（重复自动跳过）")
        self.ui.entry_product.clear()
        self.ui.entry_skc.clear()
        self.refresh_table()

    def batch_modify_skc(self):
        db, _ = get_current_database()
        if db is None:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        text, ok = QInputDialog.getText(self.ui, "批量修改 SKC", "请输入 SKC（空格隔开）:")
        if not ok or not text:
            return
        skc_list = text.strip().split()
        status, ok2 = QInputDialog.getItem(self.ui, "选择状态", "状态:", status_options, 0, False)
        if not ok2:
            return
        modified = 0
        not_found = []
        for skc in skc_list:
            s = str(skc).strip()
            found = False
            for product, skcs in db.items():
                if s in skcs:
                    skcs[s] = status
                    modified += 1
                    found = True
            if not found:
                not_found.append(s)
        save_project_data(current_project)
        self.save_database_async()
        msg = f"已修改 {modified} 个 SKC 为「{status}」"
        if not_found:
            msg += "\n未找到 SKC: " + " ".join(not_found)
        QMessageBox.information(self.ui, "完成", msg)
        self.refresh_table()

    def batch_delete_skc(self):
        db, _ = get_current_database()
        if db is None:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        text, ok = QInputDialog.getText(self.ui, "批量删除 SKC", "请输入 SKC（空格隔开）:")
        if not ok or not text:
            return
        skc_list = text.strip().split()
        deleted = 0
        not_found = []
        for skc in skc_list:
            s = str(skc).strip()
            found = False
            for product, skcs in list(db.items()):
                if s in skcs:
                    del skcs[s]
                    deleted += 1
                    found = True
            if not found:
                not_found.append(s)
        save_project_data(current_project)
        self.save_database_async()
        msg = f"已删除 {deleted} 个 SKC"
        if not_found:
            msg += "\n未找到 SKC: " + " ".join(not_found)
        QMessageBox.information(self.ui, "完成", msg)
        self.refresh_table()

    def auto_sort_by_status(self):
        db, _ = get_current_database()
        if db is None:
            return
        for product, skcs in list(db.items()):
            sorted_items = sorted(skcs.items(), key=lambda kv: status_options.index(kv[1]) if kv[1] in status_options else len(status_options))
            db[product] = dict(sorted_items)
        save_project_data(current_project)
        self.save_database_async()
        QMessageBox.information(self.ui, "完成", "已按状态顺序整理 SKC")
        self.refresh_table()

    def confirm_add_image(self):
        db, imgs = get_current_database()
        if db is None:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        product = self.ui.image_product_combo.currentText()
        if not product:
            QMessageBox.warning(self.ui, "提示", "请选择货号")
            return
        
        img_path = None
        try:
            
            img_path = self.ui.image_drop_label.image_path
        except Exception:
            img_path = None
        if not img_path:
            QMessageBox.warning(self.ui, "提示", "请拖入或粘贴图片")
            return
        imgs[product] = img_path
        save_project_data(current_project)
        self.save_database_async()
        QMessageBox.information(self.ui, "完成", f"已为货号 {product} 添加图片")
        self.refresh_table()
        self.ui.image_drop_label.clear()

    def import_excel_data(self):
        db, _ = get_current_database()
        if db is None:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        file_path, _ = QFileDialog.getOpenFileName(self.ui, "选择 Excel 文件", "", "Excel (*.xlsx *.xlsm)")
        if not file_path:
            return
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            QMessageBox.warning(self.ui, "导入失败", f"无法打开文件: {e}")
            return
        imported = 0
        for ws in wb.worksheets:
            max_col = ws.max_column
            for col in range(1, max_col+1, 2):
                product = ws.cell(row=1, column=col).value
                if not product:
                    continue
                product = str(product)
                if product not in db:
                    db[product] = {}
                max_row = ws.max_row
                for row in range(4, max_row+1):
                    skc = ws.cell(row=row, column=col).value
                    status = ws.cell(row=row, column=col+1).value
                    if skc and status:
                        s_skc = str(skc).strip()
                        if any(s_skc in skcs for p, skcs in db.items()):
                            continue
                        db[product][s_skc] = str(status)
                        imported += 1
        save_project_data(current_project)
        self.save_database_async()
        QMessageBox.information(self.ui, "完成", f"成功导入 {imported} 条记录")
        self.refresh_table()

    def open_latest_excel(self):
        if not current_project:
            QMessageBox.warning(self.ui, "提示", "请先选择项目")
            return
        open_latest_file(current_project)

    
    def on_project_changed(self, text):
        global current_project
        if text and text in projects:
            current_project = text
            self.refresh_table()

    def create_project_ui(self):
        name, ok = QInputDialog.getText(self.ui, "新建项目", "请输入项目名称（可空，默认 项目N）:")
        if not ok:
            return
        if not name:
            name = f"项目{len(projects)+1}"
        if name in projects:
            QMessageBox.warning(self.ui, "失败", "项目名已存在")
            return
        idx = len(projects) + 1
        folder = f"skc-data-{idx}"
        os.makedirs(folder, exist_ok=True)
        projects[name] = {"database": {}, "images": {}, "folder": folder}
        save_project_data(name)
        self.ui.project_combo.addItem(name)
        self.ui.project_combo.setCurrentText(name)
        QMessageBox.information(self.ui, "完成", f"已创建项目：{name}")

    def switch_project_ui(self):
        name = self.ui.project_combo.currentText()
        if not name:
            return
        if name not in projects:
            QMessageBox.warning(self.ui, "失败", "所选项目不存在")
            return
        QMessageBox.information(self.ui, "切换", f"已切换到项目：{name}")
        self.refresh_table()

    def import_project_ui(self):
        folder = QFileDialog.getExistingDirectory(self.ui, "选择要导入的项目文件夹")
        if not folder:
            return
        data_file = os.path.join(folder, "data.json")
        if not os.path.exists(data_file):
            QMessageBox.warning(self.ui, "失败", "未找到 data.json")
            return
        with open(data_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        idx = len(projects) + 1
        new_name = f"项目{idx}"
        new_folder = f"skc-data-{idx}"
        os.makedirs(new_folder, exist_ok=True)
        try:
            shutil.copy(data_file, os.path.join(new_folder, "data.json"))
        except Exception:
            with open(os.path.join(new_folder, "data.json"), "w", encoding="utf-8") as f:
                json.dump({"database": data.get("database", {}), "images": data.get("images", {})}, f, ensure_ascii=False, indent=2)
        projects[new_name] = {"database": data.get("database", {}), "images": data.get("images", {}), "folder": new_folder}
        save_project_data(new_name)
        self.ui.project_combo.addItem(new_name)
        self.ui.project_combo.setCurrentText(new_name)
        QMessageBox.information(self.ui, "导入成功", f"项目已导入为：{new_name}")
        self.refresh_table()

    def export_project_ui(self):
        global current_project
        if not current_project:
            QMessageBox.warning(self.ui, "失败", "请先选择项目")
            return
        target_folder = QFileDialog.getExistingDirectory(self.ui, "选择导出目标文件夹")
        if not target_folder:
            return
        src = projects[current_project]["folder"]
        dst = os.path.join(target_folder, current_project)
        os.makedirs(dst, exist_ok=True)
        try:
            shutil.copy(os.path.join(src, "data.json"), dst)
            for f in glob.glob(os.path.join(src, "skc_*.xlsx")):
                shutil.copy(f, dst)
            QMessageBox.information(self.ui, "导出完成", f"已导出到：{dst}")
        except Exception as e:
            QMessageBox.warning(self.ui, "导出失败", f"导出失败: {e}")

# ---------- 启动 ----------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    load_all_projects()
    ui = SKCUI()
    logic = SKCManagerLogic(ui)
    ui.show()
    sys.exit(app.exec())
